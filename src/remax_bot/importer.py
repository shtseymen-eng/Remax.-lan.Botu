from __future__ import annotations
import csv, re
from pathlib import Path
from .core import Listing, norm

FIELD_ALIASES={
 "listing_id":["ilan no","ilan numarasi","portfoy no","portfoy numarasi","id"],
 "advisor":["danisman","danisman adi","ilan sahibi","temsilci"],
 "phone":["telefon","cep telefonu","danisman telefonu","gsm"],
 "title":["baslik","ilan basligi","portfoy basligi","aciklama"],
 "price":["fiyat","ilan fiyati","bedel"],
 "location":["bolge","konum","lokasyon","il ilce mahalle","adres"],
 "rooms":["oda","oda sayisi"],
 "transaction_type":["satilik kiralik","islem tipi","durum","ilan tipi"],
 "property_type":["emlak turu","gayrimenkul turu","kategori","tip"],
 "sqm":["m2","metrekare","brut m2","net m2"],
 "listing_date":["ilan tarihi","tarih","yayin tarihi"],
 "url":["ilan linki","link","url","portfoy linki"],
 "source_url":["kaynak","ofis","kaynak linki"]
}
NOISE_ADVISOR_TOKENS={
    "portfoy","konum","kocaeli","izmit","detayli","gayrimen","re/max","remax",
    "satilik","kiralik","telefon","iletisim","ada","parsel","mahalle","ilce",
    "fiyat","aciklama","m2","metrekare","ilan no","portfoy no"
}
def _h(s): return norm(str(s or "")).replace("²","2")
def _map_headers(headers):
    out={}; nh={_h(h):h for h in headers if h is not None}
    for field,aliases in FIELD_ALIASES.items():
        for a in aliases:
            if _h(a) in nh:
                out[field]=nh[_h(a)]; break
    return out
def _row_to_listing(row,mapping,default_source="excel://import"):
    g=lambda f: str(row.get(mapping.get(f,""),"") or "").strip()
    url=g("url"); lid=g("listing_id")
    if not lid:
        m=re.search(r"(?:P)?(\d{5,})",url)
        lid=m.group(1) if m else str(abs(hash((g("title"),url,g("advisor")))))
    src=g("source_url") or default_source
    return Listing(lid,g("title") or f"İlan {lid}",url,g("advisor"),g("phone"),g("price"),g("location"),g("rooms"),g("transaction_type"),g("property_type"),g("sqm"),g("listing_date"),src)
def read_list_file(path):
    path=Path(path); ext=path.suffix.lower(); rows=[]
    if ext==".csv":
        raw=path.read_text(encoding="utf-8-sig",errors="replace")
        try: dialect=csv.Sniffer().sniff(raw[:4096],delimiters=";,\t,")
        except Exception: dialect=csv.excel
        rows=list(csv.DictReader(raw.splitlines(),dialect=dialect))
    elif ext==".xlsx":
        from openpyxl import load_workbook
        wb=load_workbook(path,read_only=True,data_only=True); ws=wb.active; vals=ws.iter_rows(values_only=True)
        headers=[str(x or "").strip() for x in next(vals,[])]; rows=[dict(zip(headers,r)) for r in vals]
    else: raise ValueError("Yalnızca .xlsx ve .csv destekleniyor.")
    if not rows:return []
    mapping=_map_headers(rows[0].keys())
    if "title" not in mapping and "url" not in mapping: raise ValueError("Başlık veya İlan Linki sütunu bulunamadı.")
    return [_row_to_listing(r,mapping) for r in rows if any(str(v or "").strip() for v in r.values())]
def price_number(s):
    digits=re.sub(r"[^0-9]","",str(s or "")); return int(digits) if digits else None
def is_valid_advisor(name):
    raw=(name or "").strip()
    if not raw or len(raw)<5 or len(raw)>70 or any(ch.isdigit() for ch in raw): return False
    n=norm(raw)
    if any(tok in n for tok in NOISE_ADVISOR_TOKENS): return False
    words=[w for w in re.split(r"\s+",raw) if w]
    if len(words)<2:return False
    letters=sum(ch.isalpha() for ch in raw)
    return letters >= max(4,int(len(raw)*0.55))
def advisor_names(items):
    names={x.advisor.strip() for x in items if is_valid_advisor(x.advisor)}
    return sorted(names,key=norm)
def filter_listings(items,advisor="",transaction="",property_type="",location="",min_price=None,max_price=None,query=""):
    rows=list(items)
    if advisor: rows=[x for x in rows if norm(x.advisor)==norm(advisor)]
    if transaction: rows=[x for x in rows if norm(transaction) in norm(x.transaction_type)]
    if property_type: rows=[x for x in rows if norm(property_type) in norm(x.property_type)]
    if location: rows=[x for x in rows if norm(location) in norm(x.location)]
    if min_price is not None: rows=[x for x in rows if price_number(x.price) is not None and price_number(x.price)>=int(min_price)]
    if max_price is not None: rows=[x for x in rows if price_number(x.price) is not None and price_number(x.price)<=int(max_price)]
    if query:
        from .core import search
        rows=search(rows,query)
    return rows
def command_help():
    return """RE/MAX ÇARŞI İLAN BOTU KOMUTLARI

#Max başla
Max ilan arama botunu aktif eder.

#Max durdur
Max ilan arama botunu durdurur.

#?
Komut listesini gösterir.

#danışmanlar
Gerçek danışmanları mevcut ilan adetleriyle listeler.

#yahya kaptan kiralık 3+1
Uygun ilanları danışman bazında sayar.

#yahya kaptan kiralık 3+1 link
Uygun ilanların linklerini danışman adıyla gönderir.

#izmit kiralık 55 bin
55.000 TL ve altındaki uygun ilanları linkleriyle gönderir.

#izmit satılık daire 4 milyon
4.000.000 TL ve altındaki uygun ilanları linkleriyle gönderir.

#danışman ayşe
Danışman adına göre ilanları bulur.

Komutun yalnızca başında # olması yeterlidir.
Eski #komut# biçimi de desteklenir.

Eski biçim örnekleri:
#danışmanlar#
#yahya kaptan kiralık 3+1#
#izmit kiralık 55 bin#
#izmit satılık daire 4 milyon#"""
def _command_body(text):
    t=(text or "").strip()
    if t=="#?": return "?",True
    if len(t)<2 or not t.startswith("#"): return "",False
    body=t[1:]
    if body.endswith("#"): body=body[:-1]
    return body.strip(),True
def parse_command(text):
    body,ok=_command_body(text)
    if not ok:return {"type":"ignore"}
    if body=="?":return {"type":"help"}
    if not body:return {"type":"ignore"}
    if norm(body)=="danismanlar":return {"type":"advisors"}
    link=bool(re.search(r"\blink\b",norm(body))); body=re.sub(r"(?i)\blink\b"," ",body).strip()
    m=re.match(r"(?i)danışman\s+(.+)$",body)
    if not m:m=re.match(r"(?i)danisman\s+(.+)$",norm(body))
    if m:return {"type":"search","query":m.group(1).strip(),"links":link,"advisor_only":True,"target_price":None}
    target=None; normalized=norm(body)
    suffix_matches=list(re.finditer(r"(?<![\d+])(\d+(?:[.,]\d+)?)\s*(bin|k|milyon|mn)(?:\s*(?:tl|₺))?(?!\w)",normalized))
    if suffix_matches:
        pm=suffix_matches[-1]; number=float(pm.group(1).replace(",","."))
        multiplier=1_000_000 if pm.group(2) in {"milyon","mn"} else 1_000
        target=int(number*multiplier)
        body=re.sub(r"(?i)(?<![\d+])\d+(?:[.,]\d+)?\s*(?:bin|k|milyon|mn)(?:\s*(?:tl|₺))?(?!\w)"," ",body,count=1).strip()
    else:
        direct=list(re.finditer(r"(?<![\d+])(?:\d{1,3}(?:[.\s]\d{3})+|\d{4,9})(?:\s*(?:tl|₺))?(?![\d+])",body,re.I))
        if direct:
            raw=direct[-1].group(0); target=int(re.sub(r"[^0-9]","",raw))
            body=(body[:direct[-1].start()]+" "+body[direct[-1].end():]).strip()
    return {"type":"search","query":body,"links":link,"target_price":target}
def command_response(items,text,limit=12):
    from .core import search
    cmd=parse_command(text)
    if cmd["type"]=="ignore":return None
    if cmd["type"]=="help":return command_help()
    if cmd["type"]=="advisors":
        by={}
        for x in items:
            if is_valid_advisor(x.advisor):
                name=x.advisor.strip(); by[name]=by.get(name,0)+1
        if not by:return "İlan havuzunda geçerli danışman bilgisi bulunamadı."
        lines=["DANIŞMAN İLAN DAĞILIMI",""]
        lines += [f"{k}: {v} ilan" for k,v in sorted(by.items(),key=lambda kv:(-kv[1],norm(kv[0])))]
        lines += ["",f"Toplam ilan: {len(items)}",f"Toplam danışman: {len(by)}"]
        return "\n".join(lines)
    matches=search(items,cmd.get("query","")); target=cmd.get("target_price")
    if target is not None:
        priced=[(price_number(x.price),x) for x in matches]; priced=[p for p in priced if p[0] is not None and p[0]<=target]
        matches=[x for _,x in sorted(priced,key=lambda p:p[0],reverse=True)]
    if not matches:
        if target is not None:return f"{target:,.0f} TL ve altında aramanıza uygun ilan bulunamadı.".replace(",",".")
        return "Aramanıza uygun ilan bulunamadı."
    if cmd.get("links") or target is not None:
        lines=[f"{target:,.0f} TL ve altında {len(matches)} uygun ilan bulundu:".replace(",",".")] if target is not None else [f"{len(matches)} uygun ilan bulundu:"]
        for x in (matches if target is not None else matches[:limit]):
            lines += ["",x.title,f"{x.price} | {x.location}",f"Danışman: {x.advisor or 'Belirtilmemiş'}",x.url or "(link yok)"]
        return "\n".join(lines)
    by={}
    for x in matches:
        name=x.advisor if is_valid_advisor(x.advisor) else "Danışman belirtilmemiş"; by[name]=by.get(name,0)+1
    return "\n".join([f"{len(matches)} uygun ilan bulundu:"]+[f"{k}: {v} ilan" for k,v in sorted(by.items(),key=lambda kv:(-kv[1],norm(kv[0])))]+["","Linkleri görmek için aynı komuta link ekleyin."])
